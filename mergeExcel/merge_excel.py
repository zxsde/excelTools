#!/usr/bin/python3

import pandas
import openpyxl
import os
from openpyxl import load_workbook
from openpyxl import Workbook


"""
使用方式：
1. 新建一个文件夹 "handle_excel"
2. 把该脚本拷贝到文件夹 "handle_excel" 下
3. 在文件夹 "handle_excel" 中再新建一个文件夹 "source"
4. 把你要处理的文件连同文件夹整个拷贝到 "source" 中

目录结构如下：
+--handle_excel
|      +--merge_excel.py
|      +--source
|      |      +--1.烟草xxx有限公司
|      |      |      +--1.1烟草xxx有限公司
|      |      |      |      +--PRC-烟草xxx有限公司.xlsx
|      |      |      |      +--PRC-烟草xxx有限公司.xlsx
|      |      |      +--1.2烟草xxx有限公司
|      |      +--2.烟草xxx有限公司
|      |      |      +--PRC-烟草xxx有限公司.xlsx
|      |      |      +--PRC-烟草xxx有限公司.xlsx
"""


#################### 配置参数

# 指定文件夹，这里包含了你所有要处理的 excel
# target_folder = "source"
target_folder = "source"

# 从每个 excel 的第几行开始合并，从 0 开始计数
offset = 2

# 要合并哪几个 sheet，从 0 开始计数，如 [7, 12] 意思是合并第 7 和 12 个 sheet
target_sheets = [
    # "表1.1-资产负债分析",
    "表1.3-其他应收款账龄分析",
    # "表1.6-其他应付款账龄分析",
    # "表1.7-预收账款账龄分析"
]

# target_file = 'target\res4.xlsx'

# 后缀，只处理该后缀的文件。目前只支持 "xlsx", "xlsm", "xls" 三种格式
suffix = ("xlsx", "xlsm", "xls")

# 前缀，不处理该前缀的文件，该前缀一般是临时文件，防止已打开的 excel 和其临时文件相互合并
prefix = ("~$", "~")


#################### 全局变量

# 所有的 excel
all_excels = []

# 合并后的数据
sheets_final = []


def traverse_folder():
    for root, dirs, files in os.walk(target_folder):
        for file in files:
            # 临时文件不统计
            if file.startswith(prefix):
                continue
            # 非 excel 不统计
            if not file.endswith(suffix):
                continue
            # 构造文件的相对路径
            file_name = os.path.join(root, file)
            all_excels.append(file_name)
            # print(file_name)
    print("扫描到 %s 个 excel 文件，请确认是否准确" % (len(all_excels)))


"""
pandas.concat默认纵向连接DataFrame对象，合并之后不改变每个DataFrame子对象的index值，横向合并可用 pandas.concat([df1, df2], axis=1)
如果两个 sheet 的列数不同，合并后以列数多的为准，短缺的列数用 NaN 填充，如果只想合并相同的列，可用 pandas.concat([df1, df2], join='inner')。
参考： [pandas中concat()的用法](https://zhuanlan.zhihu.com/p/69224745)
"""
def merge_sheet():
    # sheet name 提取一次就行，每个 excel 的 Sheetname 一样
    sheets_name = list(pandas.read_excel(all_excels[0], sheet_name=None).keys())
    print("sheet name is: %s" % sheets_name)
    for sheet_name in target_sheets:
        dfs = pandas.DataFrame()
        for file in all_excels:
            df = pandas.read_excel(file, sheet_name=sheet_name,
                                   # skiprows=offset,
                                   header=None,
                                   index_col=0
                                   )
            # print(df)
            df["source excel"] = file
            df["source sheet"] = sheet_name
            # concat默认纵向连接DataFrame对象，并且合并之后不改变每个DataFrame子对象的index值
            dfs = pandas.concat([dfs, df])


            # print(dfs)
        sheets_final.append(dfs)

# 保存到本地
def sava_file():
    with pandas.ExcelWriter(r'target\res5.xlsx') as writer:
        for i in range(len(sheets_final)):
            sheets_final[i].to_excel(writer, sheet_name=target_sheets[i])

if __name__ == '__main__':
    traverse_folder()
    merge_sheet()
    sava_file()


    # wb = load_workbook('source\1.烟草华北\1.1烟草山西\PBC-烟草山西.xlsx')
    # ws = wb["目录"]
    # outwb = Workbook()
    #
    # wb.save('balances.xlsx')


