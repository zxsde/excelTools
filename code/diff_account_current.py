#!/usr/bin/python3

import os
import sys

import shutil
import pandas
import openpyxl


# ===================================== 一般情况，仅需修改如下参数，因为每个月的文件目录/文件名都会变化

# 根路径，所有代码，excel 的所在路径
ROOT_PATH = "D:\\excelTools\\"

# 往来账款表所在路径
ACCOUNT_CURRENT_PATH = "target\\result-202104\\summary_table"

# 往来账款表
ACCOUNT_CURRENT = "往来账款.xlsx"

# Sheet 页
SHEET_NAME = "test1"

# 将要处理哪几列，因为不一定所有的列都需要进行处理，如下就是只会处理 "A,B,C" 三列
SPECIFIC_COL = "A,B,C"

# (分公司A, 分公司B, 100万) 所在的列，即公司 A、公司B、往来金额 所在列
# 注意，第几列是指 (公司 A、公司B、往来金额) 在 SPECIFIC_COL 中的第几列
SOURCE_COMPANY_A, SOURCE_COMPANY_B, SOURCE_MONEY = 1, 2, 3

# (分公司B, 分公司A, -100万) 将要保存的列，即找到的对应信息保存在哪列
TARGET_COMPANY_A, TARGET_COMPANY_B, TARGET_MONEY = "E", "F", "G"

# 往来金额所在列

# ===================================== 一般情况，仅需修改以上参数，因为每个月的文件目录/文件名都会变化


# 核对往来账款，对每项 (分公司A, 分公司B, 100万) 找到其对应的 (分公司B, 分公司A, -100万)，保存到指定的列
def diff_account_current():
    account_current_path = os.path.join(ROOT_PATH, ACCOUNT_CURRENT_PATH, ACCOUNT_CURRENT)
    print("account_current_path: \n", account_current_path, end="\n\n")
    is_exist(account_current_path)
    data = pandas.read_excel(account_current_path, sheet_name=SHEET_NAME, usecols=SPECIFIC_COL, header=None)
    # 删除包含空值的行
    data = data.dropna(axis=0, how='all')
    print(data)
    all_accounts = {}
    # 遍历表，保存所有往来账目，格式为 {(公司, 公司): (行数, 金额)}，行数是为了回写数据
    for row in data.itertuples():
        # row.Index 从 0 开始，比 excel 中实际行数少 1
        all_accounts[(row[SOURCE_COMPANY_A], row[SOURCE_COMPANY_B])] = (row.Index+1, row[SOURCE_MONEY])
    print("all_accounts", all_accounts)

    # 乙方到甲方的账款 {行数: (公司B, 公司A, 金额)}
    party_b_account = {}
    temp = {}
    # 根据甲方账款找对应的乙方账款
    for k in list(all_accounts):
        # (行数, 金额)
        v1 = all_accounts[k]
        v2 = (k[1], k[0])
        if v2 in all_accounts:
            party_b_account[v1[0]] = (k[1], k[0], all_accounts[v2][1])
            temp[v2] = (all_accounts[v2][0], all_accounts[v2][1])
            del all_accounts[k]
    print("party_b_account:", party_b_account)
    print("temp:", temp)
    for k in temp:
        if k in all_accounts:
            del all_accounts[k]
    print("can't found:", all_accounts)
    is_write = input("数据处理已完成，是否保存？(y/n):")
    if is_write == "y":
        save_data(account_current_path, party_b_account)


# 保存到本地
def save_data(account_current_path, data):
    wb = openpyxl.load_workbook(account_current_path)
    ws = wb[SHEET_NAME]
    for k, v in data.items():
        ws[TARGET_COMPANY_A+str(k)] = v[0]
        ws[TARGET_COMPANY_B+str(k)] = v[1]
        ws[TARGET_MONEY+str(k)] = v[2]
    wb.save(account_current_path)
    print("over!!!!!!!!!!!!")


# 检查文件/文件夹是否存在
def is_exist(file, is_mkdir=False, is_rm=False):
    # 文件不存在且不需要创建文件，直接退出
    if not os.path.exists(file) and not is_mkdir:
        print("file not exist: %s" % file)
        sys.exit(0)

    # 文件不存在且需要创建文件，可以递归创建目录
    elif not os.path.exists(file) and is_mkdir:
        os.makedirs(file)

    # 文件存在且需要删除文件，可以删除所有文件/文件夹
    elif os.path.exists(file) and is_rm:
        shutil.rmtree(file)


if __name__ == '__main__':
    diff_account_current()
