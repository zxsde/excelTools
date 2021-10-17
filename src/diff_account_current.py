#!/usr/bin/python3

import os

import pandas
import openpyxl

import conf.common_utils as commons_utils

"""
功能：核对往来账款差异
藐视：A列 和 B 列保存着公司、C 列保存着往来金额，经过处理后，
     A、B、C 列保存甲方到乙方发生的往来金额，E、F、G 列保存乙方到甲方发生的往来金额。
"""

# ===================================== 一般情况，仅需修改如下参数，根据实际情况进行修改

# 根路径，所有代码，excel 的所在路径
ROOT_PATH = "D:\\excelTools\\"

# 往来账款表所在路径
ACCOUNT_CURRENT_PATH = "target\\result-202109\\summary_table"

# 往来账款表
ACCOUNT_CURRENT = "merge_sheet.xlsx"

# 要处理的 Sheet 页
SHEET_NAME = "Sheet1"

# 将要处理哪几列，因为不一定所有的列都需要进行处理，如下就是只会处理 "D,E,G" 三列
SPECIFIC_COL = "D,E,G"

# (分公司A, 分公司B, 100万) 将要保存的列
ATOB_A, ATOB_B, ATOB_MONEY = "I", "J", "K"

# (分公司B, 分公司A, -100万) 将要保存的列
BTOA_A, BTOA_B, BTOA_MONEY = "M", "N", "O"

# ===================================== 一般情况，仅需修改以上参数，根据实际情况进行修改


# 核对往来账款，对每项 (分公司A, 分公司B, 100万) 找到其对应的 (分公司B, 分公司A, -100万)，保存到指定的列
def diff_account_current():
    account_current_path = os.path.join(ROOT_PATH, ACCOUNT_CURRENT_PATH, ACCOUNT_CURRENT)
    print("account_current_path: \n", account_current_path, end="\n\n")
    commons_utils.is_exist(account_current_path)
    data = pandas.read_excel(account_current_path, sheet_name=SHEET_NAME, usecols=SPECIFIC_COL)
    # 删除空行
    data = data.dropna(axis=0, how='all')
    print(data)

    all_accounts = {}  # 所有往来账目，格式为 {(公司, 公司): (行数, 金额)}
    visited = {}  # 和 b_to_a 一样，但格式为 {(公司B, 公司A): (行数, 金额)}
    a_to_b = {}  # 甲方到乙方的账款 {行数: (公司A, 公司B, 金额)}
    b_to_a = {}  # 乙方到甲方的账款 {行数: (公司B, 公司A, 金额)}

    for row in data.itertuples():
        # row.Index 从 0 开始，比 excel 中实际行数少 1，所以甲方/乙方/金额分别在第 1/2/3 列
        all_accounts[(row[1], row[2])] = (row.Index+1, row[3])
    print("all_accounts", all_accounts, end="\n\n")

    # 根据甲方账款找对应的乙方账款
    for k in list(all_accounts):
        v1 = all_accounts[k]  # (行数, 金额)
        v2 = (k[1], k[0])
        if v2 in all_accounts:
            b_to_a[v1[0]] = (k[1], k[0], all_accounts[v2][1])
            visited[v2] = (all_accounts[v2][0], all_accounts[v2][1])
            del all_accounts[k]
        if k not in visited:
            a_to_b[v1[0]] = (k[0], k[1], v1[1])
    print(" visited: %s \n a_to_b: %s \n b_to_a: %s \n" % (visited, a_to_b, b_to_a), end="\n\n")

    is_write = input("数据处理已完成，是否保存？(y/n):")
    if is_write == "y":
        save_data(account_current_path, a_to_b, b_to_a)


# 保存到本地
def save_data(account_current_path, a_to_b, b_to_a):
    wb = openpyxl.load_workbook(account_current_path)
    ws = wb[SHEET_NAME]
    # 有表头的情况下，DataFrame 中的第 0 行实际是 excel 中的第 1 行，所以要加 1
    for k, v in a_to_b.items():
        ws[ATOB_A + str(k + 1)] = v[0]
        ws[ATOB_B + str(k + 1)] = v[1]
        ws[ATOB_MONEY + str(k + 1)] = v[2]

    for k, v in b_to_a.items():
        ws[BTOA_A + str(k + 1)] = v[0]
        ws[BTOA_B + str(k + 1)] = v[1]
        ws[BTOA_MONEY + str(k + 1)] = v[2]
    wb.save(account_current_path)
    print("over!!!!!!!!!!!!")


if __name__ == '__main__':
    diff_account_current()
