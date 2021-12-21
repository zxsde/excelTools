#!/usr/bin/python3

import os
import logging

from tqdm import tqdm
from openpyxl import load_workbook

import conf.common_utils as commons_utils
import conf.constant as constant

"""
功能：检查表名
描述：检查指定文件夹下所有 excel 是否包含指定的 Sheet 页
"""


# ===================================== 一般情况，仅需修改如下参数，根据实际情况进行修改

# 根路径，所有代码，excel 的所在路径
ROOT_PATH = "E:\\excelTools\\"

# PBC 目录，所有 PBC 表所在的路径
ALL_PBC_PATH = "target\\result-202109\\all_PBC"

# 所有 excel 都应该包含如下 Sheet
STANDARD_SHEET = constant.STANDARD_SHEET

# ===================================== 一般情况，仅需修改以上参数，根据实际情况进行修改

# EXCEl 文件后缀，只处理该后缀的文件。目前不识别 xls 版本
EXCEL_SUFFIX = ("xlsx", "xlsm")

# 临时文件前缀，不处理该前缀的文件，该前缀一般是临时文件，如已打开的 excel 会额外生成一个额 ~ 开头的文件
TEMP_PREFIX = ("~$", "~")

# 所有的 excel，格式为 {excel 名: 路径+excel名}
all_pbc = {}

# 日志格式
LOG_FORMAT = "%(asctime)s - %(levelname)s - %(message)s"
logging.basicConfig(filename='..\\logs\\check_sheet_name.log', level=logging.DEBUG, format=LOG_FORMAT)


# 从 all_PBC 下获取所有的表
def get_all_pbc():
    all_pbc_path = os.path.join(ROOT_PATH, ALL_PBC_PATH)
    print("all_pbc: \n", all_pbc_path, end="\n\n")
    commons_utils.is_exist(all_pbc_path)
    for root, dirs, files in os.walk(all_pbc_path):
        for file in files:
            # 临时文件不统计
            if file.startswith(TEMP_PREFIX):
                continue
            if file.endswith("xls"):
                print(file, end="\n\n")
            # 非 excel 不统计
            if not file.endswith(EXCEL_SUFFIX):
                continue
            file_path = os.path.join(all_pbc_path, root, file)
            all_pbc[file] = file_path
            # print(file)
    print("扫描到 %s 个 excel 文件(不包含xls)，请核实：\n %s" % (len(all_pbc), all_pbc), end="\n\n")
    is_check = input("\033[1;33m 确认文件个数是否正确，是否开始检查？(y/n):")
    if is_check == "y":
        check_sheet_name()


# 检查所有 excel 是否包含指定 Sheet
def check_sheet_name():
    wrong_excel = {}  # {excel: 缺失的 Sheet}
    wrong_order = []  # Sheet 顺序不对的 excel

    # 遍历 excel
    for file, path in tqdm(all_pbc.items()):
        wb = load_workbook(path, read_only=True)
        if wb.sheetnames == STANDARD_SHEET:
            continue
        diff_set = set(STANDARD_SHEET).difference(set(wb.sheetnames))
        if diff_set:
            wrong_excel[file] = list(diff_set)
        else:
            wrong_order.append(file)

    if wrong_order:
        print("Sheet 存在但顺序不对的 excel 有 %s 个，如下，请核实： \n %s" % (len(wrong_order), wrong_order), end="\n\n")
    else:
        print("恭喜！！！所有 excel 的 Sheet 顺序也正确", end="\n\n")
    if wrong_excel:
        print("缺失 Sheet 的 excel 有 %s 个，如下，请核实： \n %s" % (len(wrong_excel), wrong_excel), end="\n\n")
        logging.debug(wrong_excel)
        is_write = input("\033[1;33m 是否对缺失 Sheet 的 excel 创建对应的空 Sheet 页(y/n):")
        if is_write:
            write_sheet(wrong_excel)
    else:
        print("恭喜！！！所有 excel 都包含指定 Sheet", end="\n\n")


# 写入空 Sheet 页
def write_sheet(wrong_excel):
    for excel, sheets in tqdm(wrong_excel.items()):
        keep_vba = False
        # xlsm 是启用了宏的文件，如果不设置keep_vba=True会造成excel无法打开
        if excel.endswith("xlsm"):
            keep_vba = True
        wb = load_workbook(all_pbc[excel], keep_vba=keep_vba)
        for sheet in sheets:
            wb.create_sheet(sheet)
        wb.save(all_pbc[excel])
    logging.debug(wrong_excel)
    print("\033[1;32m" + "Success!!!!!")


if __name__ == '__main__':
    get_all_pbc()
